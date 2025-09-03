from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

driver = webdriver.Chrome()

wb = Workbook()
ws = wb.active
ws.title = "Anúncios"

cabecalho = ['Titulo', 'Preço', 'Localização', 'Data Postada', 'Link']
ws.append(cabecalho)

for col in range(1, len(cabecalho) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

for pagina in range(1, 5):
    url = f'https://www.olx.com.br/imoveis/aluguel/estado-sp?f=p&o={pagina}'
    print(f"📄 Coletando página {pagina}...")
    driver.get(url)
    sleep(10)

    anuncios = driver.find_elements(By.XPATH, "//section[contains(@class, 'olx-adcard__horizontal')]")

    for anuncio in anuncios:
        try:
            titulo = anuncio.find_element(By.XPATH, ".//h2[contains(@class, 'olx-adcard__title')]").text
        except:
            titulo = ''
        
        try:
            preco = anuncio.find_element(By.XPATH, ".//h3[contains(@class, 'olx-adcard__price')]").text
        except:
            preco = ''

        try:
            local = anuncio.find_element(By.XPATH, ".//p[contains(@class, 'olx-adcard__location')]").text
        except:
            local = ''

        try:
            data = anuncio.find_element(By.XPATH, ".//p[contains(@class, 'olx-adcard__date')]").text
        except:
            data = ''
        
        try:
            link = anuncio.find_element(By.XPATH, ".//a[contains(@class, 'olx-adcard__link')]").get_attribute("href")
        except:
            link = ''
        
        ws.append([titulo, preco, local, data, link])

for coluna in ws.columns:
    max_length = 0
    coluna_letra = coluna[0].column_letter
    for cell in coluna:
        try:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ajuste = (max_length + 2)
    ws.column_dimensions[coluna_letra].width = ajuste

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal='left')

data_atual = datetime.now().strftime("%d-%m-%y")
nome_arquivo = f"anuncios {data_atual}.xlsx"

wb.save(nome_arquivo)

driver.quit()
print(f"✅ Arquivo '{nome_arquivo}' salvo com sucesso!")
